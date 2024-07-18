#import libraries
import face_recognition
import cv2
#brings openCV into python libraries as cv2

import numpy as np
import os

import xlwt
#This module helps in modifying and writing the data into the spreadsheets 

from xlwt import Workbook
from datetime import datetime

import xlrd, xlwt
#xlrd is for retrieving data 
from xlutils.copy import copy as xl_copy
import dlib
#for face recognition library required 
import openpyxl

#########################################################################################################################

def get_empty_columns(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    empty_columns = []

    for column in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        # Checking if all cells are empty in a given column
        if all(cell.value is None for cell in column):
            empty_columns.append(column[0].column)
    return empty_columns
####################################################################################################################

def search_and_write_value(sheet, search_value,column_Number):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value == search_value:
                sheet.cell(row=cell.row, column=column_Number, value=1)
    wb.save(fileName)
#################################################################################################################


CurrentFolder = os.getcwd()
image1 = CurrentFolder+'\\SudhaMurty.jpg'
image2 = CurrentFolder+'\\RatanTata.jpg'
image3 = CurrentFolder+ '\\NarayanaMurthy.jpg'


#0 is for laptop webcam
video_capture = cv2.VideoCapture(0)


person1_name = "SudhaMurty"
person1_image = face_recognition.load_image_file(image1)
person1_face_encoding = face_recognition.face_encodings(person1_image)[0]
#A face encoding is basically a way to represent the face using a set of 128 computer-generated measurements

person2_name = "RatanTata"
person2_image = face_recognition.load_image_file(image2)
person2_face_encoding = face_recognition.face_encodings(person2_image)[0]

person3_name = "NarayanaMurthy"
person3_image = face_recognition.load_image_file(image3)
person3_face_encoding = face_recognition.face_encodings(person3_image)[0]

known_face_encodings = [ person1_face_encoding, person2_face_encoding,person3_face_encoding]
known_face_names = [person1_name, person2_name, person3_name]


#now we begin checking out the faces
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True


already_attendence_taken = ""
#so far no name, no student's attendance is taken, initialising

fileName = 'Teacher.xlsx'
sheetName = input("Please enter the subject and class: ")
wb = openpyxl.load_workbook(fileName)
sheet = wb[sheetName]
empty_columns = get_empty_columns(fileName,sheetName)
column_Number = empty_columns[0]


current_date = datetime.now().date()
sheet.cell(row = 1, column = empty_columns[0], value = current_date)


while True:
            # capturing a single frame of the video
            ret, frame = video_capture.read()

            # reducing the size to 1/4th (0.25) for faster processing
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

            
            #OpenCV i.e. is cv2 uses BGR ---> RGB for face_recognition library
            rgb_small_frame = small_frame[:, :, ::-1]

            
            if process_this_frame:
                # We want to find all faces in a single frame
                #multiple studnets in a classroom can be detected at once
                face_locations = face_recognition.face_locations(rgb_small_frame)
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations, model="small")

                face_names = []
                for face_encoding in face_encodings:
                    matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                    name = "Unknown"
                    
                    # from the known face encodings, we try to find teh closest face
                    face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]:
                        name = known_face_names[best_match_index]

                    face_names.append(name)
                    if((already_attendence_taken != name) and (name != "Unknown")):
                        search_and_write_value(sheet,name,column_Number)
                        print("attendence taken")
                        #we mark it as attendance taken
                        already_attendence_taken = name                       
   
                    else:
                     print("next student")
                        
            process_this_frame = not process_this_frame
            #we are processing alternate frames


            #show a rectangle around the face in the live video
            for (top, right, bottom, left), name in zip(face_locations, face_names):
                #we had reduced it to 1/4th size, so scale it back up
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4

                
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

                
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

            # Display the resulting image
            cv2.imshow('Video', frame)

            # we hit q for qutting the frame
            if cv2.waitKey(1) & 0xff==ord('q'):   
                print("data save")
                break


video_capture.release()     
cv2.destroyAllWindows()

